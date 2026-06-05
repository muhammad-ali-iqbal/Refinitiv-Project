"""
ingestion/parser.py

Parses a single Refinitiv CF Export .xlsx file into a normalised DataFrame.

Output shape:
    ric | company_name | country | industry | sheet | section | variable | year | value | currency | scaling
"""

import re
import datetime
import pandas as pd
from openpyxl import load_workbook

# Sheets with a different structure (segment breakdowns) — skip for now
SKIP_SHEETS = {"Business Line By Segm", "Business Line By Stat", "Major Customer", "Pension"}

# Metadata rows at the top of every sheet — stop parsing when we hit "Field Name"
META_KEYS = {
    "Company Name", "Country of Exchange", "Country of Headquarters",
    "TRBC Industry Group", "CF Template", "Consolidation Basis",
    "Scaling", "Period", "Export Date", "Statement Data",
    "Period End Date", "Statement Date", "Standardized Currency",
    "Template Type", "Measure system",
}

# Section header detection: rows with a value in col A but no numeric data
def _is_section_header(row_values: list) -> bool:
    """A section header has text in col A and nothing numeric in the rest."""
    if not row_values[0]:
        return False
    data_cols = row_values[1:]
    return all(v is None or str(v).strip() == "" for v in data_cols)


def _parse_ric(company_name_cell: str) -> tuple[str, str]:
    """
    'Pakistan State Oil Company Ltd (PSO.PSX)' 
    → ('Pakistan State Oil Company Ltd', 'PSO.PSX')
    """
    match = re.search(r"\(([^)]+)\)\s*$", str(company_name_cell))
    if match:
        ric = match.group(1)
        name = company_name_cell[: company_name_cell.rfind("(")].strip()
        return name, ric
    return str(company_name_cell).strip(), ""


def parse_file(filepath: str) -> pd.DataFrame:
    """
    Parse all relevant sheets in a CF Export file.
    Returns a tidy DataFrame with one row per (company, sheet, variable, year).
    """
    wb = load_workbook(filepath, read_only=True, data_only=True)
    all_rows = []

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            continue

        # ── Extract metadata ──────────────────────────────────────────────────
        meta = {}
        field_name_row_idx = None

        for i, row in enumerate(rows):
            if not row or row[0] is None:
                continue
            key = str(row[0]).strip()
            if key in META_KEYS:
                meta[key] = row[1] if len(row) > 1 else None
            if key == "Field Name":
                field_name_row_idx = i
                # Year columns are the remaining cells in this row.
                # openpyxl returns date-formatted cells as datetime objects;
                # extract just the year integer to avoid "2020-12-31 00:00:00" strings.
                def _extract_year_label(v):
                    if not v:
                        return None
                    if isinstance(v, (datetime.datetime, datetime.date)):
                        return str(v.year)
                    s = str(v).strip()
                    # Also handle strings that look like dates (e.g. "2020-12-31")
                    try:
                        parsed = pd.to_datetime(s, dayfirst=True, errors="raise")
                        return str(parsed.year)
                    except Exception:
                        return s
                year_cols = [_extract_year_label(v) for v in row[1:]]
                break

        if field_name_row_idx is None:
            continue  # sheet has no data table

        company_raw = meta.get("Company Name", "")
        company_name, ric = _parse_ric(company_raw)
        currency = meta.get("Standardized Currency", "")
        scaling = meta.get("Scaling", "")

        # currency/scaling may be repeated across columns — take first non-None
        if isinstance(currency, str) is False:
            currency = ""
        if isinstance(scaling, str) is False:
            scaling = ""

        # ── Parse data rows ───────────────────────────────────────────────────
        current_section = ""

        for row in rows[field_name_row_idx + 1 :]:
            if not row or row[0] is None:
                continue

            variable = str(row[0]).strip()
            if not variable:
                continue

            # Skip known meta-like stragglers
            if variable in META_KEYS:
                continue

            data_values = list(row[1:])

            if _is_section_header([variable] + data_values):
                current_section = variable
                continue

            # Emit one row per year
            for col_idx, year_label in enumerate(year_cols):
                if not year_label:
                    continue
                if col_idx >= len(data_values):
                    break
                val = data_values[col_idx]
                if val is None or str(val).strip() == "":
                    continue
                try:
                    val = float(val)
                except (ValueError, TypeError):
                    continue  # skip non-numeric cells

                all_rows.append({
                    "ric":          ric,
                    "company_name": company_name,
                    "sheet":        sheet_name,
                    "section":      current_section,
                    "variable":     variable,
                    "year":         year_label,
                    "value":        val,
                    "currency":     currency,
                    "scaling":      scaling,
                })

    wb.close()

    if not all_rows:
        return pd.DataFrame()

    return pd.DataFrame(all_rows)
