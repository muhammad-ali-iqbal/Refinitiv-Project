"""
pipeline/excel_writer.py — Writes one .xlsx workbook per company.

Four sheets: Income Statement | Balance Sheet | Cash Flow | Ratios
The Ratios sheet uses live Excel formulas referencing the three statement sheets.
"""

import os
import logging
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from pipeline.fetcher import STATEMENTS

log = logging.getLogger(__name__)

# ── Colour palette ─────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F3864")   # Dark navy
SUBHEAD_FILL  = PatternFill("solid", fgColor="2E75B6")   # Blue
ALT_ROW_FILL  = PatternFill("solid", fgColor="EBF3FB")   # Light blue
WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")

HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
LABEL_FONT    = Font(name="Calibri", bold=False, size=10)
DATA_FONT     = Font(name="Calibri", size=10, color="1F497D")  # Blue = raw data
FORMULA_FONT  = Font(name="Calibri", size=10, color="375623")  # Green = formula

THIN = Side(style="thin", color="BDD7EE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COL_LABEL_WIDTH = 32
COL_DATA_WIDTH  = 14


def _apply_header(ws, row: int, col: int, value: str, span: int = 1) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def _apply_data_cell(ws, row: int, col: int, value, fmt: str, is_formula=False) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = FORMULA_FONT if is_formula else DATA_FONT
    cell.number_format = fmt
    cell.alignment = Alignment(horizontal="right")
    cell.border = BORDER


def _write_statement_sheet(ws, df: pd.DataFrame, fields: list[tuple], title: str) -> None:
    """Write a single financial statement to a worksheet."""
    fmt_map = {f[1]: f[2] for f in fields}  # label → Excel format

    if df.empty:
        ws.cell(row=1, column=1, value="No data available for this company.")
        return

    years = list(df.columns)
    n_years = len(years)

    # Row 1: sheet title
    ws.cell(row=1, column=1, value=title).font = Font(name="Calibri", bold=True, size=12)

    # Row 2: blank

    # Row 3: headers
    _apply_header(ws, 3, 1, "Line Item")
    for i, yr in enumerate(years, start=2):
        _apply_header(ws, 3, i, str(yr)[:4])   # show 4-char year

    # Row 4+: data
    for r_idx, (label, row_data) in enumerate(df.iterrows(), start=4):
        fill = ALT_ROW_FILL if r_idx % 2 == 0 else WHITE_FILL
        label_cell = ws.cell(row=r_idx, column=1, value=label)
        label_cell.font = LABEL_FONT
        label_cell.fill = fill
        label_cell.border = BORDER

        fmt = fmt_map.get(str(label), "#,##0")
        for c_idx, yr in enumerate(years, start=2):
            val = row_data.get(yr)
            val = None if pd.isna(val) else val
            data_cell = ws.cell(row=r_idx, column=c_idx, value=val)
            data_cell.font = DATA_FONT
            data_cell.fill = fill
            data_cell.number_format = fmt
            data_cell.alignment = Alignment(horizontal="right")
            data_cell.border = BORDER

    # Column widths
    ws.column_dimensions["A"].width = COL_LABEL_WIDTH
    for i in range(2, n_years + 2):
        ws.column_dimensions[get_column_letter(i)].width = COL_DATA_WIDTH


def _write_ratios_sheet(ws, statements_data: dict, years: list) -> None:
    """
    Write a Ratios sheet with live Excel formulas cross-referencing
    the three statement sheets.
    """
    n_years = len(years)

    # Row lookup helpers — find the Excel row number for a label in a sheet
    def find_row(sheet_name: str, label: str) -> int | None:
        sheet = ws.parent[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == label:
                    return cell.row
        return None

    rev_row    = find_row("Income Statement", "Total Revenue")
    cor_row    = find_row("Income Statement", "Cost of Revenue")
    ebitda_row = find_row("Income Statement", "EBITDA")
    ni_row     = find_row("Income Statement", "Net Income")
    debt_row   = find_row("Balance Sheet",    "Total Debt")
    eq_row     = find_row("Balance Sheet",    "Total Equity")
    cash_row   = find_row("Balance Sheet",    "Cash & ST Investments")
    assets_row = find_row("Balance Sheet",    "Total Assets")
    capex_row  = find_row("Cash Flow",        "Capital Expenditures")
    fcf_row    = find_row("Cash Flow",        "Free Cash Flow")

    def _ok(*rows):
        return all(r is not None for r in rows)

    # Each ratio: (label, formula_template, required_row_vars)
    RATIOS = [
        ("Profitability", None, None),
        ("Gross Margin %",    "=IFERROR(('Income Statement'!{rev_row}{c}-'Income Statement'!{cor_row}{c})/'Income Statement'!{rev_row}{c},\"\")", _ok(rev_row, cor_row)),
        ("EBITDA Margin %",   "=IFERROR('Income Statement'!{ebitda_row}{c}/'Income Statement'!{rev_row}{c},\"\")",                               _ok(ebitda_row, rev_row)),
        ("Net Margin %",      "=IFERROR('Income Statement'!{ni_row}{c}/'Income Statement'!{rev_row}{c},\"\")",                                   _ok(ni_row, rev_row)),
        ("Leverage", None, None),
        ("Debt / Equity",     "=IFERROR('Balance Sheet'!{debt_row}{c}/'Balance Sheet'!{eq_row}{c},\"\")",                                        _ok(debt_row, eq_row)),
        ("Net Debt / EBITDA", "=IFERROR(('Balance Sheet'!{debt_row}{c}-'Balance Sheet'!{cash_row}{c})/'Income Statement'!{ebitda_row}{c},\"\")", _ok(debt_row, cash_row, ebitda_row)),
        ("Efficiency", None, None),
        ("Asset Turnover",    "=IFERROR('Income Statement'!{rev_row}{c}/'Balance Sheet'!{assets_row}{c},\"\")",                                  _ok(rev_row, assets_row)),
        ("CapEx / Revenue",   "=IFERROR(ABS('Cash Flow'!{capex_row}{c})/'Income Statement'!{rev_row}{c},\"\")",                                  _ok(capex_row, rev_row)),
        ("Cash Flow", None, None),
        ("FCF Conversion",    "=IFERROR('Cash Flow'!{fcf_row}{c}/'Income Statement'!{ni_row}{c},\"\")",                                          _ok(fcf_row, ni_row)),
        ("FCF Margin %",      "=IFERROR('Cash Flow'!{fcf_row}{c}/'Income Statement'!{rev_row}{c},\"\")",                                         _ok(fcf_row, rev_row)),
    ]

    ws.cell(row=1, column=1, value="Ratios").font = Font(name="Calibri", bold=True, size=12)
    _apply_header(ws, 3, 1, "Ratio")
    for i, yr in enumerate(years, start=2):
        _apply_header(ws, 3, i, str(yr)[:4])

    r = 4
    for label, formula_template, deps_ok in RATIOS:
        if formula_template is None:
            # Section header
            cell = ws.cell(row=r, column=1, value=label)
            cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
            cell.fill = SUBHEAD_FILL
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_years + 1)
            r += 1
            continue

        fill = ALT_ROW_FILL if r % 2 == 0 else WHITE_FILL
        label_cell = ws.cell(row=r, column=1, value=label)
        label_cell.font = LABEL_FONT
        label_cell.fill = fill
        label_cell.border = BORDER

        for c_idx, _ in enumerate(years, start=2):
            col_letter = get_column_letter(c_idx)

            if deps_ok:
                formula = formula_template.format(
                    c=col_letter,
                    rev_row=rev_row, cor_row=cor_row, ebitda_row=ebitda_row,
                    ni_row=ni_row, debt_row=debt_row, eq_row=eq_row,
                    cash_row=cash_row, assets_row=assets_row,
                    capex_row=capex_row, fcf_row=fcf_row,
                )
            else:
                formula = "N/A"

            cell = ws.cell(row=r, column=c_idx, value=formula)
            cell.font = FORMULA_FONT
            cell.fill = fill
            cell.number_format = "0.0%"
            cell.alignment = Alignment(horizontal="right")
            cell.border = BORDER

        r += 1

    ws.column_dimensions["A"].width = COL_LABEL_WIDTH
    for i in range(2, n_years + 2):
        ws.column_dimensions[get_column_letter(i)].width = COL_DATA_WIDTH


def write_workbook(
    output_path: str,
    company_name: str,
    statements_data: dict,
) -> None:
    """
    Write a 4-sheet workbook for one company.

    Args:
        output_path: Full path to the .xlsx file to create.
        company_name: Used only for logging.
        statements_data: Dict returned by fetcher.get_financials().
    """
    wb = Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    # Determine year columns from whichever statement has data
    years = []
    for df in statements_data.values():
        if not df.empty:
            years = list(df.columns)
            break

    for stmt_name, fields in STATEMENTS.items():
        ws = wb.create_sheet(title=stmt_name)
        df = statements_data.get(stmt_name, pd.DataFrame())
        _write_statement_sheet(ws, df, fields, stmt_name)

    if years:
        ws_ratios = wb.create_sheet(title="Ratios")
        _write_ratios_sheet(ws_ratios, statements_data, years)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    log.info("Saved: %s", output_path)
